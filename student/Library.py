import os
from datetime import datetime

import openpyxl

class Common:
    def __init__(self, FileNamePath, SheetName):
        self.vk = openpyxl.load_workbook(FileNamePath)
        self.sh = self.vk[SheetName]

    def fetch_row_count(self):
        return self.sh.max_row

    def fetch_column_count(self):
        return self.sh.max_column

    def fetch_key_name(self):
        c = self.sh.max_column
        li = []
        for i in range(1, c + 1):
            cell = self.sh.cell(row=1, column=i)
            li.append(cell.value)  # Use append instead of insert
        return li

    def update_request_with_data(self, rowNumber, jsonRequest, keyList):
        c = self.sh.max_column
        for i in range(1, c + 1):
            cell = self.sh.cell(row=rowNumber, column=i)
            jsonRequest[keyList[i - 1]] = cell.value
        return jsonRequest

    def format_date(self, date):
        """Convert a datetime object to a string in the format 'YYYY-MM-DD'."""
        if isinstance(date, datetime):
            return date.strftime("%Y-%m-%d")
        return date

    def construct_file_path(self, *path_segments):
        current_dir = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(current_dir, *path_segments)